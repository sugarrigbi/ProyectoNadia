from App.Models.Case_Models import (
    Caso as Tabla_Caso, 
    Usuario as Tabla_Usuario, 
    Estado_Caso as Tabla_Estado, 
    Prioridad as Tabla_Prioridad, 
    Incidente as Tabla_Incidente, 
    Barrio as Tabla_Barrio, 
    Localidad as Tabla_Localidad, 
    Ciudad as Tabla_Ciudad, 
    Departamento as Tabla_Dep, 
    Tipo_Relacion as Tabla_Relacion, 
    Caso_Discusion as Tabla_Discusion, 
    Departamento as T_Departamento, 
    Ciudad as T_Ciudad, 
    Localidad as T_Localidad, 
    Barrio as T_Barrio, 
    Casos_a_Casos as Tabla_Casos_a_Casos, 
    Radicado_Caso as Tabla_Radicado, 
    Caso_Auditoria as Tabla_Auditoria, 
    RolAPermiso as Tabla_Permiso
)
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from App.Utilities.Tables import db
from openpyxl import Workbook
from datetime import datetime
from sqlalchemy import func
import requests
import json
import os
import io

class Case_Service():
    @staticmethod
    def Create(Data_C, Data_L, User_ID, Data_C_C=None, Data_R=None):
        if not User_ID:
            return "Auth"
        
        Usuario_Validar = Tabla_Usuario.query.get(User_ID)
        Permisos = Tabla_Permiso.query.filter(Tabla_Permiso.Rol_ID == Usuario_Validar.Rol_ID).all()
        Nombres = [p.to_dict()["Nombre"] for p in Permisos]
        if "caso_crear" not in Nombres and "caso_crear_propio" not in Nombres:
            return "Auth"

        if Data_L:
            if Data_L["Departamento_ID"]:
                Departamento = T_Departamento.query.get(int(Data_L["Departamento_ID"]))
            else:
                Departamento_Norm = Data_L["Departamento"].capitalize().strip()
                Departamento = T_Departamento.query.filter_by(Nombre=Departamento_Norm).first()
                if not Departamento:
                    Departamento = T_Departamento(Nombre=Departamento_Norm, Pais_ID=1)
                    db.session.add(Departamento)
                    db.session.flush()
            if Data_L["Ciudad_ID"]:
                Ciudad = T_Ciudad.query.get(int(Data_L["Ciudad_ID"]))
            else:
                Ciudad_Norm = Data_L["Ciudad"].capitalize().strip()
                Ciudad = T_Ciudad.query.filter_by(Nombre=Ciudad_Norm).first()
                if not Ciudad:
                    if Data_L["Departamento_ID"]:
                        Ciudad = T_Ciudad(Nombre=Ciudad_Norm, Departamento_ID=int(Data_L["Departamento_ID"]))
                        db.session.add(Ciudad)
                        db.session.flush()
                    else:
                        Ciudad = T_Ciudad(Nombre=Ciudad_Norm, Departamento_ID=Departamento.ID)
                        db.session.add(Ciudad)
                        db.session.flush()
            if Data_L["Localidad_ID"]:
                Localidad = T_Localidad.query.get(int(Data_L["Localidad_ID"]))       
            else:
                Localidad_Norm = Data_L["Localidad"].capitalize().strip()
                Localidad = T_Localidad.query.filter_by(Nombre=Localidad_Norm).first()
                if not Localidad:
                    if Data_L["Ciudad_ID"]:
                        Localidad = T_Localidad(Nombre=Localidad_Norm, Ciudad_ID=int(Data_L["Ciudad_ID"]))
                        db.session.add(Localidad)
                        db.session.flush()                        
                    else:
                        Localidad = T_Localidad(Nombre=Localidad_Norm, Ciudad_ID=Ciudad.ID)  
                        db.session.add(Localidad)
                        db.session.flush()   
            if Data_L["Barrio_ID"]:
                Barrio = T_Barrio.query.get(int(Data_L["Barrio_ID"]))
            else:
                Barrio_Norm = Data_L["Barrio"].capitalize().strip()
                Barrio = T_Barrio.query.filter_by(Nombre=Barrio_Norm).first()
                if not Barrio:
                    if Data_L["Localidad_ID"]:
                        Barrio = T_Barrio(Nombre=Barrio_Norm, Localidad_ID=int(Data_L["Localidad_ID"]))
                        db.session.add(Barrio)
                        db.session.flush()
                    else:
                        Barrio = T_Barrio(Nombre=Barrio_Norm, Localidad_ID=Localidad.ID)
                        db.session.add(Barrio)
                        db.session.flush()             

        if "caso_crear_propio" in Nombres:
            if Data_C:
                Data_C["Barrio_ID"] = Barrio.ID
                Data_C["Nombre"] = "Caso No Gestionado"
                Data_C["Usuario_Creador_ID"] = Usuario_Validar.ID
                Data_C["Estado_Caso_ID"] = 1
                Data_C["Prioridad_ID"] = 1
                Data_C["Usuario_Asociado_ID"] = 1
                Data_C["Actualizado_En"] = datetime.now()
                Caso = Tabla_Caso(**Data_C)
                db.session.add(Caso)
                db.session.flush()            

                Radicados = Tabla_Radicado.query.order_by(Tabla_Radicado.ID.desc()).first()
                if not Radicados:
                    Radicado_Final = Tabla_Radicado(Radicado="000001R", Caso_ID=Caso.ID)
                else:
                    Radicado_text = int(Radicados.Radicado.rstrip("R"))+1
                    Radicado_com = f"{Radicado_text:06d}R"          
                    Radicado_Final = Tabla_Radicado(Radicado=Radicado_com, Caso_ID=Caso.ID)
                db.session.add(Radicado_Final)
                db.session.flush()   

        elif "caso_crear" in Nombres:
            if Data_C:
                Data_C["Barrio_ID"] = Barrio.ID
                Data_C["Actualizado_En"] = datetime.now()
                Caso = Tabla_Caso(**Data_C)
                db.session.add(Caso)
                db.session.flush()            

                Radicados = Tabla_Radicado.query.order_by(Tabla_Radicado.ID.desc()).first()
                Radicado_text = int(Radicados.Radicado.rstrip("R"))+1
                Radicado_com = f"{Radicado_text:06d}R"          
                Radicado_Final = Tabla_Radicado(Radicado=Radicado_com, Caso_ID=Caso.ID)
                db.session.add(Radicado_Final)
                db.session.flush() 

            if Data_R.get("Relacion_Radicado"):
                Caso_a_Relacionar = Tabla_Caso.query.filter_by(ID=Data_R["Relacion_Radicado"]).first()
                if not Caso_a_Relacionar:
                    return False
                else:
                    Relacion = Tabla_Casos_a_Casos(Caso_Principal_ID=Caso.ID,Caso_Asociado_ID=int(Data_R["Relacion_Radicado"]), Tipo_Relacion_ID=int(Data_R["Relacion_Tipo"]))
                    db.session.add(Relacion)
                    db.session.flush()

            if Data_C_C["Mensaje"]:
                Data_C_C["Caso_ID"] = Caso.ID
                Caso_Discucion = Tabla_Discusion(**Data_C_C)
                db.session.add(Caso_Discucion)
                db.session.flush()

        Caso_json = Caso.to_dict(include_relationships=True)
        Caso_json_text = json.dumps(Caso_json, default=str, ensure_ascii=False)

        Auditoria = Tabla_Auditoria(Accion="Caso creado", Anterior=Caso_json_text, Modificado_Por=Usuario_Validar.ID, Caso_ID=Caso.ID)
        db.session.add(Auditoria)
        db.session.commit()

        requests.post(os.getenv("EMAIL_SERVICE"),
            json={
                "Template": "Caso_Creado",
                "Datos": {
                    "Nombre": Caso_json["usuario_creador"]["Persona"]["Primer_Nombre"]+" "+Caso_json["usuario_creador"]["Persona"]["Primer_Apellido"], 
                    "Radicado": Caso_json["radicados"][0]["Radicado"], 
                    },
                "Correo": Caso_json["usuario_creador"]["Correo"],
                "Asunto": f"Caso {Caso_json['radicados'][0]['Radicado']} creado"
            }
        )
        return Caso
    @staticmethod
    def Read_All(User_ID, Pagina):
        if not User_ID:
            return "Auth"
        
        Usuario_Validar = Tabla_Usuario.query.get(User_ID)
        Permisos = Tabla_Permiso.query.filter(Tabla_Permiso.Rol_ID == Usuario_Validar.Rol_ID).all()
        Nombres = [p.to_dict()["Nombre"] for p in Permisos]          
        if "caso_ver_propio" in Nombres:
            Querry = Tabla_Caso.query.filter(Tabla_Caso.Estado_Caso_ID != 4, Tabla_Caso.Usuario_Creador_ID == Usuario_Validar.ID).order_by(Tabla_Caso.ID.asc())
            Casos = Querry.paginate(page=Pagina, per_page=int(os.getenv("CASE_PER_PAGE")),error_out=False)
            if Pagina < 1 or Pagina > Casos.pages:
                Pagina = Casos.pages
                Casos = Querry.paginate(page=Pagina, per_page=int(os.getenv("CASE_PER_PAGE")),error_out=False)
            return Casos
        elif "caso_ver" in Nombres:   
            Querry = Tabla_Caso.query.filter(Tabla_Caso.Estado_Caso_ID != 4).order_by(Tabla_Caso.ID.desc())
            Casos = Querry.paginate(page=Pagina, per_page=int(os.getenv("CASE_PER_PAGE")),error_out=False)        
            if Pagina < 1 or Pagina > Casos.pages:
                Pagina = Casos.pages
                Casos = Querry.paginate(page=Pagina, per_page=int(os.getenv("CASE_PER_PAGE")),error_out=False)
            return Casos
        else:
            return "Auth"           
    @staticmethod
    def Obtener_Datos(User_ID):
        if not User_ID:
            return "Auth"
        
        Usuario_Validar = Tabla_Usuario.query.get(User_ID)
        Permisos = Tabla_Permiso.query.filter(Tabla_Permiso.Rol_ID == Usuario_Validar.Rol_ID).all()
        Nombres = [p.to_dict()["Nombre"] for p in Permisos]            
        if "caso_ver" in Nombres or "caso_ver_propio" in Nombres:
            Usuarios = Tabla_Usuario.query.order_by(Tabla_Usuario.ID.asc()).all()
            Estados = Tabla_Estado.query.filter(Tabla_Estado.ID != 4).order_by(Tabla_Estado.ID.asc()).all()
            Prioridades = Tabla_Prioridad.query.order_by(Tabla_Prioridad.ID.asc()).all()
            Incidentes = Tabla_Incidente.query.order_by(Tabla_Incidente.ID.asc()).all()
            Barrios = Tabla_Barrio.query.order_by(Tabla_Barrio.ID.asc()).all()
            Localidades = Tabla_Localidad.query.order_by(Tabla_Localidad.ID.asc()).all()
            Ciudades = Tabla_Ciudad.query.order_by(Tabla_Ciudad.ID.asc()).all()
            Departamentos = Tabla_Dep.query.order_by(Tabla_Dep.ID.asc()).all()
            Relaciones = Tabla_Relacion.query.filter(Tabla_Relacion.ID != 4).order_by(Tabla_Relacion.ID.asc()).all()          
            Data = {
                "Estados": [e.to_dict() for e in Estados],
                "Prioridades": [p.to_dict() for p in Prioridades],
                "Incidentes": [i.to_dict() for i in Incidentes],
                "Barrios": [b.to_dict() for b in Barrios],
                "Localidades": [l.to_dict() for l in Localidades],
                "Ciudades": [c.to_dict() for c in Ciudades],
                "Departamentos": [d.to_dict() for d in Departamentos],
                "Relaciones": [r.to_dict() for r in Relaciones]
            }
            if "caso_ver" in Nombres:
                Data["Usuarios"] = [u.to_dict() for u in Usuarios]
            elif "caso_ver_propio" in Nombres:
                Data["Usuarios"] = [u.to_dict2() for u in Usuarios]
            else:
                Data["Usuarios"] = []
            return Data    
        else:
            return "Auth"       
    @staticmethod
    def Linea_Tiempo(User_ID, Casos):
        if not User_ID:
            return "Auth"
        
        Auditoria = {}
        Usuario_Validar = Tabla_Usuario.query.get(User_ID)
        Permisos = Tabla_Permiso.query.filter(Tabla_Permiso.Rol_ID == Usuario_Validar.Rol_ID).all()
        Nombres = [p.to_dict()["Nombre"] for p in Permisos]
        if "caso_ver_propio" in Nombres or "caso_ver" in Nombres:
            Casos_json = {C.ID for C in Casos.items}
            Auditoria = Tabla_Auditoria.query.filter(Tabla_Auditoria.Caso_ID.in_(Casos_json)).order_by(Tabla_Auditoria.ID.asc()).all() 
            if not Auditoria:
                return False  
            return Auditoria                    
        else:
            return "Auth"  
    @staticmethod
    def Read_By(Filtros, User_ID, Pagina):
        if not User_ID:
            return "Auth"
        
        Usuario_Validar = Tabla_Usuario.query.get(User_ID)
        Permisos = Tabla_Permiso.query.filter(Tabla_Permiso.Rol_ID == Usuario_Validar.Rol_ID).all()
        if "caso_ver" not in [p.to_dict()["Nombre"] for p in Permisos]:
            return "Auth"

        Casos = Tabla_Caso.query
        if Filtros.get("Estado"):
            Casos = Casos.filter(Tabla_Caso.Estado_Caso_ID.in_(Filtros["Estado"]))
        if Filtros.get("Usuario_Encargado"):
            Casos = Casos.filter(Tabla_Caso.Usuario_Asociado_ID.in_(Filtros["Usuario_Encargado"]))
        if Filtros.get("Usuario_Creador"):
            Casos = Casos.filter(Tabla_Caso.Usuario_Creador_ID.in_(Filtros["Usuario_Creador"]))
        if Filtros.get("Prioridad"):
            Casos = Casos.filter(Tabla_Caso.Prioridad_ID.in_(Filtros["Prioridad"]))            
        if Filtros.get("Incidente"):
            Casos = Casos.filter(Tabla_Caso.Incidente_ID.in_(Filtros["Incidente"]))
        if Filtros.get("Nombre"):
            nombre = Filtros['Nombre'][0]         
            Casos = Casos.filter(Tabla_Caso.Nombre.ilike(f"%{nombre}%"))

        Paginacion = Casos.order_by(Tabla_Caso.ID.asc()).paginate(page=Pagina, per_page=int(os.getenv("CASE_PER_PAGE")),error_out=False)
        if Pagina < 1 or Pagina > Paginacion.pages:
            Pagina = Paginacion.pages
            Paginacion = Casos.order_by(Tabla_Caso.ID.asc()).paginate(page=Pagina, per_page=int(os.getenv("CASE_PER_PAGE")),error_out=False)

        return Paginacion
    @staticmethod
    def Update(Case_ID, Data_C, Data_C_C, Data_L, Data_R, User_ID):
        if not User_ID:
            return "Auth"
        
        Usuario_Validar = Tabla_Usuario.query.get(User_ID)
        Permisos = Tabla_Permiso.query.filter(Tabla_Permiso.Rol_ID == Usuario_Validar.Rol_ID).all()
        if "caso_editar" not in [p.to_dict()["Nombre"] for p in Permisos]:
            return "Auth"            

        Caso = Tabla_Caso.query.get(Case_ID)
        Caso_json = Caso.to_dict(include_relationships=True)
        Caso_json_text = json.dumps(Caso_json, default=str, ensure_ascii=False)
        if not Caso:
            return False
        
        if Data_L:
            Departamento = Data_L["Departamento"].capitalize().strip()
            Dep_Exi = T_Departamento.query.filter(T_Departamento.Nombre == Departamento, T_Departamento.Pais_ID == 1).first()
            if not Dep_Exi:
                Dep_Exi = T_Departamento(Nombre=Departamento, Pais_ID=1)
                db.session.add(Dep_Exi)
                db.session.flush()

            Ciudad = Data_L["Ciudad"].capitalize().strip()
            Ciu_Exi = T_Ciudad.query.filter(T_Ciudad.Nombre == Ciudad, T_Ciudad.Departamento_ID == Dep_Exi.ID).first()
            if not Ciu_Exi:
                Ciu_Exi = T_Ciudad(Nombre=Ciudad, Departamento_ID=Dep_Exi.ID)
                db.session.add(Ciu_Exi)
                db.session.flush()                

            Localidad = Data_L["Localidad"].capitalize().strip()
            Loc_Exi = T_Localidad.query.filter(T_Localidad.Nombre == Localidad, T_Localidad.Ciudad_ID == Ciu_Exi.ID).first()
            if not Loc_Exi:
                Loc_Exi = T_Localidad(Nombre=Localidad, Ciudad_ID=Ciu_Exi.ID)
                db.session.add(Loc_Exi)
                db.session.flush()

            Barrio = Data_L["Barrio"].capitalize().strip()
            Bar_Exi = T_Barrio.query.filter(T_Barrio.Nombre == Barrio, T_Barrio.Localidad_ID == Loc_Exi.ID).first()
            if not Bar_Exi:
                Bar_Exi = T_Barrio(Nombre=Barrio, Localidad_ID=Loc_Exi.ID)
                db.session.add(Bar_Exi)
                db.session.flush()

        Estado_Correo = Tabla_Estado.query.get(int(Data_C['Estado_Caso_ID']))
        if not Estado_Correo:
            return False

        nuevo_estado = int(Data_C['Estado_Caso_ID'])
        anterior_estado = Caso.Estado_Caso_ID

        if Data_C:
            Caso.Creacion = Data_C['Creacion']
            Caso.Nombre = Data_C['Nombre']
            Caso.Descripcion = Data_C['Descripcion']
            Caso.Afectados = Data_C['Afectados']
            Caso.Direccion = Data_C['Direccion']
            Caso.Usuario_Creador_ID = int(Data_C['Usuario_Creador_ID'])
            Caso.Usuario_Asociado_ID = int(Data_C['Usuario_Asociado_ID'])
            Caso.Incidente_ID = int(Data_C['Incidente_ID'])
            Caso.Estado_Caso_ID = int(Data_C['Estado_Caso_ID'])
            Caso.Prioridad_ID = int(Data_C['Prioridad_ID'])
            Caso.Creacion = Data_C['Creacion']
            Caso.Barrio_ID = Bar_Exi.ID
            db.session.flush()

        if Data_C_C["Mensaje"]:
            Caso_Discucion = Tabla_Discusion(**Data_C_C)
            db.session.add(Caso_Discucion)
            db.session.flush()

        if Data_R.get("Relacion_Radicado"):
            Caso_a_Relacionar = Tabla_Caso.query.filter_by(ID=Data_R["Relacion_Radicado"]).first()
            if not Caso_a_Relacionar:
                return False
            else:
                Relacion = Tabla_Casos_a_Casos(Caso_Principal_ID=Caso.ID,Caso_Asociado_ID=int(Data_R["Relacion_Radicado"]), Tipo_Relacion_ID=int(Data_R["Relacion_Tipo"]))
                db.session.add(Relacion)
                db.session.flush()

        Auditoria = Tabla_Auditoria(Accion="Caso modificado", Anterior=Caso_json_text, Modificado_Por=Data_C_C["Usuario_ID"], Caso_ID=Case_ID)
        db.session.add(Auditoria)
        db.session.commit()

        if nuevo_estado == 3 and nuevo_estado != anterior_estado:
            requests.post(os.getenv("EMAIL_SERVICE"),
                json={
                    "Template": "Caso_Resuelto",
                    "Datos": {
                        "Nombre": Caso_json["usuario_creador"]["Persona"]["Primer_Nombre"]+" "+Caso_json["usuario_creador"]["Persona"]["Primer_Apellido"], 
                        "Radicado": Caso_json["radicados"][0]["Radicado"], 
                        "Estado": Estado_Correo.Nombre},
                    "Correo": Caso_json["usuario_creador"]["Correo"],
                    "Asunto": f"Caso {Caso_json['radicados'][0]['Radicado']} resuelto"
                }
            )    
        elif nuevo_estado != 3:
            if nuevo_estado != anterior_estado:
                requests.post(os.getenv("EMAIL_SERVICE"),
                    json={
                        "Template": "Caso_Cambio",
                        "Datos": {
                            "Nombre": Caso_json["usuario_creador"]["Persona"]["Primer_Nombre"]+" "+Caso_json["usuario_creador"]["Persona"]["Primer_Apellido"], 
                            "Radicado": Caso_json["radicados"][0]["Radicado"], 
                            "Estado": Estado_Correo.Nombre},
                        "Correo": Caso_json["usuario_creador"]["Correo"],
                        "Asunto": "Cambio en el estado del caso"
                    }
                )        

        return Caso
    @staticmethod
    def Delete(Case_ID, User_ID, User_ID2):
        if not User_ID2:
            return "Auth"
        
        Usuario_Validar = Tabla_Usuario.query.get(User_ID2)
        Permisos = Tabla_Permiso.query.filter(Tabla_Permiso.Rol_ID == Usuario_Validar.Rol_ID).all()
        if "caso_eliminar" not in [p.to_dict()["Nombre"] for p in Permisos]:
            return "Auth"

        Caso = Tabla_Caso.query.get(Case_ID)
        Caso_json_text = json.dumps(Caso.to_dict(include_relationships=True), default=str, ensure_ascii=False)

        if not Caso:
            return False
        
        Caso.Estado_Caso_ID = 4

        Auditoria = Tabla_Auditoria(Accion="Caso eliminado", Anterior=Caso_json_text, Modificado_Por=User_ID, Caso_ID=Caso.ID)
        db.session.add(Auditoria)
        db.session.commit()

        return Caso
    @staticmethod
    def Delete_Relacion(CasoHijo_Rad, CasePadre_Rad, Tipo_Relacion, User_ID):
        Caso_Padre = Tabla_Radicado.query.filter(Tabla_Radicado.Radicado == CasePadre_Rad).first()
        Caso_Hijo = Tabla_Radicado.query.filter(Tabla_Radicado.Radicado == CasoHijo_Rad).first()

        Relacion = Tabla_Casos_a_Casos.query.filter(Tabla_Casos_a_Casos.Caso_Principal_ID == Caso_Padre.ID).filter(Tabla_Casos_a_Casos.Caso_Asociado_ID == Caso_Hijo.ID).filter(Tabla_Casos_a_Casos.Tipo_Relacion_ID == Tipo_Relacion).first()
        Relacion2 = Relacion.to_dict2()
        Relacion2_text = json.dumps(Relacion2, default=str, ensure_ascii=False)
        Auditoria = Tabla_Auditoria(Accion="Relacion Eliminada", Anterior=Relacion2_text, Modificado_Por=User_ID, Caso_ID=Caso_Padre.ID)
        db.session.add(Auditoria)        
        db.session.delete(Relacion)

        db.session.commit()        
        return Caso_Padre
    @staticmethod
    def Obtener_Estadisticas(User_ID):
        if not User_ID:
            return "Auth"
        
        Usuario_Validar = Tabla_Usuario.query.get(User_ID)
        Permisos = Tabla_Permiso.query.filter(Tabla_Permiso.Rol_ID == Usuario_Validar.Rol_ID).all()
        Nombres = [p.to_dict()["Nombre"] for p in Permisos]          
        if "estadisticas_ver" in Nombres:
            Estados = db.session.query(Tabla_Estado.Nombre, func.count(Tabla_Caso.ID)).join(Tabla_Caso, Tabla_Caso.Estado_Caso_ID == Tabla_Estado.ID).group_by(Tabla_Estado.Nombre).all()
            Prioridades = db.session.query(Tabla_Prioridad.Prioridad, func.count(Tabla_Caso.ID)).join(Tabla_Caso, Tabla_Caso.Prioridad_ID == Tabla_Prioridad.ID).group_by(Tabla_Prioridad.ID, Tabla_Prioridad.Prioridad).order_by(Tabla_Prioridad.ID).all()
            Incidentes = db.session.query(Tabla_Incidente.Incidente, func.count(Tabla_Caso.ID)).join(Tabla_Caso, Tabla_Caso.Incidente_ID == Tabla_Incidente.ID).group_by(Tabla_Incidente.Incidente).all()
            Usuarios = db.session.query(Tabla_Usuario.Nombre, func.count(Tabla_Caso.ID)).join(Tabla_Caso, Tabla_Caso.Usuario_Asociado_ID == Tabla_Usuario.ID).group_by(Tabla_Usuario.Nombre).order_by(func.count(Tabla_Caso.ID).desc()).limit(9).all()
            Tendencia_Total = db.session.query(func.date_format(Tabla_Caso.Creacion, "%Y-%m"), func.count(Tabla_Caso.ID)).group_by(func.date_format(Tabla_Caso.Creacion, "%Y-%m")).order_by(func.date_format(Tabla_Caso.Creacion, "%Y-%m")).all()
            Tendencia_Resuelto = db.session.query(func.date_format(Tabla_Caso.Creacion, "%Y-%m"), func.count(Tabla_Caso.ID)).filter(Tabla_Caso.Estado_Caso_ID == 3).group_by(func.date_format(Tabla_Caso.Creacion, "%Y-%m")).order_by(func.date_format(Tabla_Caso.Creacion, "%Y-%m")).all()
            Tendencia_Final = {}
            for Mes, Conteo in Tendencia_Total:
                Tendencia_Final[Mes] = {"Creados": Conteo, "Resueltos": 0}
            for Mes, Conteo in Tendencia_Resuelto:
                if Mes in Tendencia_Final:
                    Tendencia_Final[Mes]["Resueltos"] = Conteo

            Data = {
                "Estados": {Nombre: Conteo for Nombre, Conteo in Estados},
                "Prioridades": {Nombre: Conteo for Nombre, Conteo in Prioridades},
                "Incidentes": {Nombre: Conteo for Nombre, Conteo in Incidentes},
                "Usuarios": {Nombre: Conteo for Nombre, Conteo in Usuarios},
                "Tendencia_Final": Tendencia_Final
            }
            return Data
        else:
            return "Auth"
    @staticmethod
    def Exportar_Exel(User_ID):
        if not User_ID:
            return "Auth"
        
        Usuario_Validar = Tabla_Usuario.query.get(User_ID)
        Permisos = Tabla_Permiso.query.filter(Tabla_Permiso.Rol_ID == Usuario_Validar.Rol_ID).all()
        Nombres = [P.to_dict()["Nombre"] for P in Permisos]
        if "exportar_excel" in Nombres:
            Casos = Tabla_Caso.query.filter(Tabla_Caso.Estado_Caso_ID != 4).order_by(Tabla_Caso.ID.asc()).all()
            if not Casos:
                return None

            Libro = Workbook()
            Hoja = Libro.active
            Hoja.title = "Reporte de casos"
            Hoja.row_dimensions[1].height = 25
            Hoja.freeze_panes = "A2"

            Color_Principal = "5B2D8E"
            Color_Filas = "F4EEFF"
            Color_Bordes = "DDDDDD"
            
            Estado_Colores = {
                "Pendiente": "FFF1B8",
                "Activo": "CFF5D2",
                "Resuelto": "BFE9F2",
                "Eliminado": "F8C7C7",
                "En espera del usuario": "E6E6E6",
                "Escalado a supervisor": "CFE0FF",
                "Reabierto": "D9D9D9",
                "Tomando desicion": "E3D6FF",
                "En espera del asesor": "EAEAEA"
            }
            Prioridad_Colores = {
                "Muy Baja": "DDE1E6",
                "Baja": "BFE9F2",
                "Media": "FFE08A",
                "Alta": "FF9E9E",
                "Critica": "C62828"
            }
            Incidente_Colores = {
                "Desplazamiento": "CFE0FF",
                "Predios Despojados": "E3D6FF",
                "Expropiacion": "FFE6A7",
                "Hurto": "FFB3B3"
            }

            Fuente_Header = Font(bold=True, color="FFFFFF", size=12)
            Fondo_Header = PatternFill(fill_type="solid", fgColor=Color_Principal)
            Fondo_Filas = PatternFill(fill_type="solid", fgColor=Color_Filas)
            Alineacion_Central = Alignment(horizontal="center", vertical="center")
            Alineacion_Start = Alignment(horizontal="left", vertical="center")
            Bordes = Border(
                left = Side(style="thin", color=Color_Bordes),
                right = Side(style="thin", color=Color_Bordes),
                top = Side(style="thin", color=Color_Bordes),
                bottom = Side(style="thin", color=Color_Bordes)
            )

            Columnas_Archivo = [
                {
                    "Titulo": "Radicado", 
                    "Ancho": 14,
                    "Valor": lambda Caso: Caso.radicados[0].Radicado
                },
                {
                    "Titulo": "Nombre", 
                    "Ancho": 35,
                    "Valor": lambda Caso: Caso.Nombre
                },
                {
                    "Titulo": "Descripcion", 
                    "Ancho": 45,
                    "Valor": lambda Caso: Caso.Descripcion
                },
                {
                    "Titulo": "Estado", 
                    "Ancho": 22,
                    "Valor": lambda Caso: Caso.estado.Nombre
                },
                {
                    "Titulo": "Prioridad", 
                    "Ancho": 14,
                    "Valor": lambda Caso: Caso.prioridad.Prioridad
                },                
                {
                    "Titulo": "Incidente", 
                    "Ancho": 20,
                    "Valor": lambda Caso: Caso.incidente.Incidente
                },                                
                {
                    "Titulo": "Afectados", 
                    "Ancho": 14,
                    "Valor": lambda Caso: Caso.Afectados
                },
                {
                    "Titulo": "Direccion", 
                    "Ancho": 25,
                    "Valor": lambda Caso: Caso.Direccion
                },
                {
                    "Titulo": "Fecha_Creacion", 
                    "Ancho": 22,
                    "Valor": lambda Caso: Caso.Creacion
                },
                {
                    "Titulo": "Usuario_Creador_ID", 
                    "Ancho": 22,
                    "Valor": lambda Caso: Caso.usuario_creador.ID
                },
                {
                    "Titulo": "Usuario_Creador_Nombre", 
                    "Ancho": 30,
                    "Valor": lambda Caso: Caso.usuario_creador.Nombre
                },
                {
                    "Titulo": "Usuario_Encargado_ID", 
                    "Ancho": 25,
                    "Valor": lambda Caso: Caso.usuario_asociado.ID
                },
                {
                    "Titulo": "Usuario_Encargado_Nombre", 
                    "Ancho": 32,
                    "Valor": lambda Caso: Caso.usuario_asociado.Nombre
                },
                {
                    "Titulo": "Barrio",
                    "Ancho": 20,
                    "Valor": lambda Caso: Caso.barrio.Nombre
                },
                {
                    "Titulo": "Localidad",
                    "Ancho": 20,
                    "Valor": lambda Caso: Caso.barrio.localidad.Nombre
                },
                {
                    "Titulo": "Ciudad",
                    "Ancho": 20,
                    "Valor": lambda Caso: Caso.barrio.localidad.ciudad.Nombre
                },
                {
                    "Titulo": "Departamento",
                    "Ancho": 22,
                    "Valor": lambda Caso: Caso.barrio.localidad.ciudad.departamento.Nombre
                }
            ]

            for Numero_Columna, Columna in enumerate(Columnas_Archivo, start=1):
                Celda = Hoja.cell(row=1, column=Numero_Columna, value=Columna["Titulo"])
                Celda.font = Fuente_Header
                Celda.fill = Fondo_Header
                Celda.alignment = Alineacion_Central
                Celda.border = Bordes   

                Letra_de_Columna = get_column_letter(Numero_Columna)
                Hoja.column_dimensions[Letra_de_Columna].width = Columna["Ancho"]

            for Numero_Fila, Caso in enumerate(Casos, start=2):
                Hoja.row_dimensions[Numero_Fila].height = 22
                for Numero_Columna, Columna in enumerate(Columnas_Archivo, start=1):
                    Valor = Columna["Valor"](Caso)
                    Celda = Hoja.cell(row=Numero_Fila, column=Numero_Columna, value=Valor)
                    Celda.border = Bordes
                    if Columna["Titulo"] == "Radicado":
                        Celda.alignment = Alineacion_Central
                        Celda.font = Font(bold=True)
                    elif Columna["Titulo"] in ["Estado", "Prioridad", "Incidente", "Afectados", "Fecha_Creacion", "Usuario_Creador_ID", "Usuario_Creador_Nombre", "Usuario_Encargado_ID", "Usuario_Encargado_Nombre"]:
                        Celda.alignment = Alineacion_Central
                    else:
                        Celda.alignment = Alineacion_Start
                    if Numero_Fila % 2 != 0:
                        Celda.fill = Fondo_Filas

                    if Columna["Titulo"] == "Estado":
                        Celda.fill = PatternFill(fill_type="solid", fgColor=Estado_Colores.get(Celda.value, "FFFFFF"))
                    elif Columna["Titulo"] == "Prioridad":
                        Celda.fill = PatternFill(fill_type="solid", fgColor=Prioridad_Colores.get(Celda.value, "FFFFFF"))
                    elif Columna["Titulo"] == "Incidente":
                        Celda.fill = PatternFill(fill_type="solid", fgColor=Incidente_Colores.get(Celda.value, "FFFFFF"))

            Hoja.auto_filter.ref = f"A1:{get_column_letter(Hoja.max_column)}1"
            Archivo = io.BytesIO()
            Libro.save(Archivo)
            Archivo.seek(0)

            return Archivo     
        else:
            return "Auth"